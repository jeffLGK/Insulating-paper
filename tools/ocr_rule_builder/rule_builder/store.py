"""把確認後的規則寫入 OCR 規則 xlsx。"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Side

# 對應現有 7 欄結構
COLS = [
    "目錄名稱", "廠牌輸出名稱", "圖上廠牌關鍵字", "圖上型號文字",
    "型號輸出", "忽略字串", "特殊規則說明",
]
_thin = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def append_rule_row(xlsx_path: Path, row: dict) -> int:
    """row 鍵：dir_name, brand, ocr_brand, ocr_model, model_out, ignore, special。
    回傳寫入的列號。"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    r = ws.max_row + 1
    values = [
        row.get("dir_name", ""), row.get("brand", ""),
        row.get("ocr_brand", ""), row.get("ocr_model", ""),
        row.get("model_out", ""), row.get("ignore", ""),
        row.get("special", ""),
    ]
    for c, v in enumerate(values, 1):
        cell = ws.cell(r, c, v if v else None)
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    wb.save(xlsx_path)
    return r
